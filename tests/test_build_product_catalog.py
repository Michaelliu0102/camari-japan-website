from __future__ import annotations

import json
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from scripts.build_product_catalog import build_catalog, read_rows, write_template


class BuildProductCatalogTests(unittest.TestCase):
    def test_write_template_creates_expected_sheets(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            workbook_path = Path(tmp_dir) / "template.xlsx"
            write_template(workbook_path)

            workbook = load_workbook(workbook_path)
            self.assertIn("product_types", workbook.sheetnames)
            self.assertIn("skus", workbook.sheetnames)
            self.assertEqual(workbook["product_types"]["A2"].value, "alcantara-panel")

    def test_write_template_uses_local_upload_paths_for_image_examples(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            workbook_path = Path(tmp_dir) / "template.xlsx"
            write_template(workbook_path)

            workbook = load_workbook(workbook_path)
            product_type_headers = [cell.value for cell in workbook["product_types"][1]]
            sku_headers = [cell.value for cell in workbook["skus"][1]]
            product_type_seo_image_column = product_type_headers.index("seo_image") + 1
            sku_image_column = sku_headers.index("image") + 1
            sku_swatch_image_column = sku_headers.index("swatch_image") + 1

            self.assertEqual(workbook["product_types"].cell(row=2, column=product_type_seo_image_column).value, "/uploads/alcantara/panel/alcantara-panel-hero.jpg")
            self.assertEqual(workbook["skus"].cell(row=2, column=sku_image_column).value, "/uploads/alcantara/panel/c-alc-4991-main.jpg")
            self.assertEqual(workbook["skus"].cell(row=2, column=sku_swatch_image_column).value, "/uploads/alcantara/swatches/c-alc-4991-swatch.jpg")
            self.assertEqual(workbook["sku_case_gallery"]["C2"].value, "/uploads/alcantara/cases/c-alc-4991-installed.jpg")

    def test_build_catalog_parses_workbook_into_product_types_and_skus(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            workbook_path = Path(tmp_dir) / "catalog.xlsx"
            write_template(workbook_path)

            workbook = load_workbook(workbook_path)
            workbook["product_type_specs"].append(
                ["alcantara-panel", 2, "width", "WIDTH", "幅", "width", "142 cm", "142 cm"]
            )
            workbook["skus"].append(
                [
                    "c-alc-735b-umber-brown",
                    "alcantara",
                    "alcantara-panel",
                    "C-ALC-735B",
                    "Umber Brown",
                    "アンバーブラウン",
                    "#735B33",
                    "https://example.com/umber-main.jpg",
                    "",
                    "A warm brown shade.",
                    "温かみのあるブラウン。",
                    "C-ALC-735B Umber Brown | CAMARI JAPAN",
                    "C-ALC-735B アンバーブラウン | CAMARI JAPAN",
                    "View Umber Brown Alcantara details.",
                    "アンバーブラウン Alcantara の詳細。",
                    "https://example.com/umber-main.jpg",
                ]
            )
            workbook["sku_specs"].append(
                ["c-alc-735b-umber-brown", 1, "Thickness", "厚み", "0.95 mm", "0.95 mm"]
            )
            workbook.save(workbook_path)

            catalog = build_catalog(read_rows(workbook_path))

            self.assertEqual(len(catalog["productTypes"]), 1)
            self.assertEqual(catalog["productTypes"][0]["specTemplate"][1]["key"], "width")
            self.assertEqual(len(catalog["skus"]), 2)
            self.assertEqual(catalog["skus"][1]["productTypeSlug"], "alcantara-panel")
            self.assertEqual(catalog["skus"][1]["specs"][0]["value"]["en"], "0.95 mm")

    def test_build_catalog_rejects_missing_product_type_reference(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            workbook_path = Path(tmp_dir) / "catalog.xlsx"
            write_template(workbook_path)

            workbook = load_workbook(workbook_path)
            workbook["skus"].append(
                [
                    "bad-sku",
                    "alcantara",
                    "missing-type",
                    "BAD-001",
                    "Bad",
                    "不正",
                    "#000000",
                    "https://example.com/bad.jpg",
                    "",
                    "Bad data",
                    "不正データ",
                    "Bad",
                    "不正",
                    "Bad",
                    "不正",
                    "https://example.com/bad.jpg",
                ]
            )
            workbook.save(workbook_path)

            with self.assertRaisesRegex(ValueError, "references missing product type"):
                build_catalog(read_rows(workbook_path))


if __name__ == "__main__":
    unittest.main()
