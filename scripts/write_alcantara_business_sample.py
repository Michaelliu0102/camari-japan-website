#!/usr/bin/env python3

from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

ROOT = Path(__file__).resolve().parent.parent
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from scripts.build_product_catalog import SHEETS


OUTPUT_PATH = ROOT / "data/import/alcantara-business-sample.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="2F6690")
SUBTLE_FILL = PatternFill("solid", fgColor="EAF1F7")
HEADER_FONT = Font(color="FFFFFF", bold=True)
BODY_FONT = Font(color="1F2937", size=10)
THIN_BORDER = Border(
    left=Side(style="thin", color="D0D7DE"),
    right=Side(style="thin", color="D0D7DE"),
    top=Side(style="thin", color="D0D7DE"),
    bottom=Side(style="thin", color="D0D7DE"),
)

DATA = {
    "product_types": [
        {
            "product_type_slug": "alcantara-panel",
            "material_slug": "alcantara",
            "name_en": "Alcantara Panel",
            "name_ja": "Alcantara パネル",
            "seo_title_en": "Alcantara Panel for Interior Surfaces | CAMARI JAPAN",
            "seo_title_ja": "Alcantara パネル 内装サーフェス向け | CAMARI JAPAN",
            "seo_description_en": "Architectural and mobility panel application data for Alcantara, including specifications, care, and technical downloads.",
            "seo_description_ja": "Alcantara パネル用途の仕様、ケア情報、技術資料をまとめたサンプルデータです。",
            "seo_image": "/uploads/alcantara/panel/alcantara-panel-hero.jpg",
        },
        {
            "product_type_slug": "alcantara-cover",
            "material_slug": "alcantara",
            "name_en": "Alcantara Cover",
            "name_ja": "Alcantara カバー",
            "seo_title_en": "Alcantara Cover for Seating and Trim | CAMARI JAPAN",
            "seo_title_ja": "Alcantara カバー シート・加飾向け | CAMARI JAPAN",
            "seo_description_en": "Covering application data for Alcantara used on seating, armrests, and soft-touch trim surfaces.",
            "seo_description_ja": "シート、アームレスト、ソフトトリム向け Alcantara カバー用途のサンプルデータです。",
            "seo_image": "/uploads/alcantara/cover/alcantara-cover-hero.jpg",
        },
    ],
    "product_type_specs": [
        {
            "product_type_slug": "alcantara-panel",
            "sort_order": 1,
            "spec_key": "application",
            "label_en": "APPLICATION",
            "label_ja": "用途",
            "aliases": "application|use",
            "default_value_en": "Wall panel / door trim / instrument garnish",
            "default_value_ja": "壁面パネル / ドアトリム / インパネ加飾",
        },
        {
            "product_type_slug": "alcantara-panel",
            "sort_order": 2,
            "spec_key": "thickness",
            "label_en": "THICKNESS",
            "label_ja": "厚み",
            "aliases": "thickness|gauge",
            "default_value_en": "0.95 mm",
            "default_value_ja": "0.95 mm",
        },
        {
            "product_type_slug": "alcantara-panel",
            "sort_order": 3,
            "spec_key": "width",
            "label_en": "WIDTH",
            "label_ja": "幅",
            "aliases": "width",
            "default_value_en": "142 cm",
            "default_value_ja": "142 cm",
        },
        {
            "product_type_slug": "alcantara-panel",
            "sort_order": 4,
            "spec_key": "backing",
            "label_en": "BACKING",
            "label_ja": "裏加工",
            "aliases": "backing|support",
            "default_value_en": "Adhesive lamination compatible",
            "default_value_ja": "接着ラミネート対応",
        },
        {
            "product_type_slug": "alcantara-cover",
            "sort_order": 1,
            "spec_key": "application",
            "label_en": "APPLICATION",
            "label_ja": "用途",
            "aliases": "application|use",
            "default_value_en": "Seat cover / armrest / console wrap",
            "default_value_ja": "シートカバー / アームレスト / コンソール巻き",
        },
        {
            "product_type_slug": "alcantara-cover",
            "sort_order": 2,
            "spec_key": "thickness",
            "label_en": "THICKNESS",
            "label_ja": "厚み",
            "aliases": "thickness|gauge",
            "default_value_en": "1.20 mm",
            "default_value_ja": "1.20 mm",
        },
        {
            "product_type_slug": "alcantara-cover",
            "sort_order": 3,
            "spec_key": "width",
            "label_en": "WIDTH",
            "label_ja": "幅",
            "aliases": "width",
            "default_value_en": "142 cm",
            "default_value_ja": "142 cm",
        },
        {
            "product_type_slug": "alcantara-cover",
            "sort_order": 4,
            "spec_key": "touch",
            "label_en": "HAND FEEL",
            "label_ja": "触感",
            "aliases": "touch|handfeel",
            "default_value_en": "Soft-touch suede finish",
            "default_value_ja": "ソフトタッチのスエード調",
        },
    ],
    "product_type_certifications": [
        {
            "product_type_slug": "alcantara-panel",
            "sort_order": 1,
            "text_en": "Carbon neutral production program",
            "text_ja": "カーボンニュートラル生産プログラム",
        },
        {
            "product_type_slug": "alcantara-panel",
            "sort_order": 2,
            "text_en": "RoHS / REACH aligned material control",
            "text_ja": "RoHS / REACH に準拠した材料管理",
        },
        {
            "product_type_slug": "alcantara-cover",
            "sort_order": 1,
            "text_en": "Abrasion performance suitable for seating programs",
            "text_ja": "着座用途に対応する耐摩耗性能",
        },
        {
            "product_type_slug": "alcantara-cover",
            "sort_order": 2,
            "text_en": "Low-VOC interior application support",
            "text_ja": "低 VOC 内装用途サポート",
        },
    ],
    "product_type_maintenance": [
        {
            "product_type_slug": "alcantara-panel",
            "sort_order": 1,
            "title_en": "Daily Care",
            "title_ja": "日常メンテナンス",
            "description_en": "Wipe with a clean damp microfiber cloth and dry immediately after cleaning.",
            "description_ja": "清潔なマイクロファイバークロスを軽く湿らせて拭き、仕上げに乾拭きしてください。",
        },
        {
            "product_type_slug": "alcantara-cover",
            "sort_order": 1,
            "title_en": "Seating Care",
            "title_ja": "シート向けケア",
            "description_en": "Use neutral cleaner for light soiling and avoid strong solvent-based products.",
            "description_ja": "軽い汚れは中性洗剤で対応し、強い溶剤系クリーナーは避けてください。",
        },
    ],
    "skus": [
        {
            "sku_slug": "c-alc-4991-shadow-black-panel",
            "material_slug": "alcantara",
            "product_type_slug": "alcantara-panel",
            "code": "C-ALC-4991-PNL",
            "color_name_en": "Shadow Black",
            "color_name_ja": "シャドウブラック",
            "hex": "#1B1B1D",
            "image": "/uploads/alcantara/panel/c-alc-4991-main.jpg",
            "swatch_image": "/uploads/alcantara/swatches/c-alc-4991-swatch.jpg",
            "summary_en": "Deep black panel finish for premium door inserts, IP garnish, and architectural trim surfaces.",
            "summary_ja": "プレミアムなドアインサート、インパネ加飾、建材トリム向けの深みあるブラック仕上げ。",
            "seo_title_en": "C-ALC-4991-PNL Shadow Black Panel | CAMARI JAPAN",
            "seo_title_ja": "C-ALC-4991-PNL シャドウブラック パネル | CAMARI JAPAN",
            "seo_description_en": "Shadow Black Alcantara panel sample with specification and technical downloads.",
            "seo_description_ja": "Shadow Black の Alcantara パネル向けサンプル仕様と技術資料です。",
            "seo_image": "/uploads/alcantara/panel/c-alc-4991-main.jpg",
        },
        {
            "sku_slug": "c-alc-6206-saddle-tan-cover",
            "material_slug": "alcantara",
            "product_type_slug": "alcantara-cover",
            "code": "C-ALC-6206-CVR",
            "color_name_en": "Saddle Tan",
            "color_name_ja": "サドルタン",
            "hex": "#9B6A3C",
            "image": "/uploads/alcantara/cover/c-alc-6206-main.jpg",
            "swatch_image": "/uploads/alcantara/swatches/c-alc-6206-swatch.jpg",
            "summary_en": "Warm tan cover material tuned for seating programs, armrests, and stitched trim applications.",
            "summary_ja": "シート、アームレスト、ステッチ加飾用途に向く温かみのあるタン系カバー材。",
            "seo_title_en": "C-ALC-6206-CVR Saddle Tan Cover | CAMARI JAPAN",
            "seo_title_ja": "C-ALC-6206-CVR サドルタン カバー | CAMARI JAPAN",
            "seo_description_en": "Saddle Tan Alcantara cover sample with care guide and application data.",
            "seo_description_ja": "Saddle Tan の Alcantara カバー向けサンプル仕様とケア情報です。",
            "seo_image": "/uploads/alcantara/cover/c-alc-6206-main.jpg",
        },
    ],
    "sku_specs": [
        {
            "sku_slug": "c-alc-4991-shadow-black-panel",
            "sort_order": 1,
            "label_en": "Thickness",
            "label_ja": "厚み",
            "value_en": "0.95 mm",
            "value_ja": "0.95 mm",
        },
        {
            "sku_slug": "c-alc-4991-shadow-black-panel",
            "sort_order": 2,
            "label_en": "Width",
            "label_ja": "幅",
            "value_en": "142 cm",
            "value_ja": "142 cm",
        },
        {
            "sku_slug": "c-alc-4991-shadow-black-panel",
            "sort_order": 3,
            "label_en": "Recommended Use",
            "label_ja": "推奨用途",
            "value_en": "Door trim / panel insert / display plinth",
            "value_ja": "ドアトリム / パネルインサート / ディスプレイ什器",
        },
        {
            "sku_slug": "c-alc-4991-shadow-black-panel",
            "sort_order": 4,
            "label_en": "MOQ",
            "label_ja": "最小ロット",
            "value_en": "1 roll trial / project quote",
            "value_ja": "試作 1 ロール / 量産は別途見積",
        },
        {
            "sku_slug": "c-alc-6206-saddle-tan-cover",
            "sort_order": 1,
            "label_en": "Thickness",
            "label_ja": "厚み",
            "value_en": "1.20 mm",
            "value_ja": "1.20 mm",
        },
        {
            "sku_slug": "c-alc-6206-saddle-tan-cover",
            "sort_order": 2,
            "label_en": "Width",
            "label_ja": "幅",
            "value_en": "142 cm",
            "value_ja": "142 cm",
        },
        {
            "sku_slug": "c-alc-6206-saddle-tan-cover",
            "sort_order": 3,
            "label_en": "Recommended Use",
            "label_ja": "推奨用途",
            "value_en": "Seat center / armrest / console wrap",
            "value_ja": "シートセンター / アームレスト / コンソール巻き",
        },
        {
            "sku_slug": "c-alc-6206-saddle-tan-cover",
            "sort_order": 4,
            "label_en": "MOQ",
            "label_ja": "最小ロット",
            "value_en": "Sampling yardage available",
            "value_ja": "カットサンプル対応可",
        },
    ],
    "sku_downloads": [
        {
            "sku_slug": "c-alc-4991-shadow-black-panel",
            "sort_order": 1,
            "title_en": "Technical Data Sheet",
            "title_ja": "技術データシート",
            "description_en": "Composition, width, thickness, and lamination notes for panel applications.",
            "description_ja": "パネル用途向けの組成、幅、厚み、ラミネート適性に関する資料。",
            "href": "/catalogs/alcantara-panel-shadow-black-tds.pdf",
            "type": "technical",
        },
        {
            "sku_slug": "c-alc-4991-shadow-black-panel",
            "sort_order": 2,
            "title_en": "Care Guide",
            "title_ja": "ケアガイド",
            "description_en": "Recommended maintenance routine for installed panel surfaces.",
            "description_ja": "施工済みパネル面の推奨メンテナンス手順。",
            "href": "/catalogs/alcantara-panel-care-guide.pdf",
            "type": "care",
        },
        {
            "sku_slug": "c-alc-6206-saddle-tan-cover",
            "sort_order": 1,
            "title_en": "Technical Data Sheet",
            "title_ja": "技術データシート",
            "description_en": "Cover application specifications including softness, backing, and sewing suitability.",
            "description_ja": "カバー用途向けの触感、裏加工、縫製適性を含む仕様資料。",
            "href": "/catalogs/alcantara-cover-saddle-tan-tds.pdf",
            "type": "technical",
        },
        {
            "sku_slug": "c-alc-6206-saddle-tan-cover",
            "sort_order": 2,
            "title_en": "Program Catalog Excerpt",
            "title_ja": "掲載カタログ抜粋",
            "description_en": "Sample catalog spread showing seating and trim colorway usage.",
            "description_ja": "シートおよびトリム用途の色展開を示すカタログ抜粋。",
            "href": "/catalogs/alcantara-cover-program-catalog.pdf",
            "type": "catalog",
        },
    ],
    "sku_case_gallery": [
        {
            "sku_slug": "c-alc-4991-shadow-black-panel",
            "sort_order": 1,
            "image": "/uploads/alcantara/cases/c-alc-4991-installed.jpg",
            "alt_en": "Shadow Black panel installed on a premium door trim insert",
            "alt_ja": "プレミアムドアトリムに施工した Shadow Black パネル",
        },
        {
            "sku_slug": "c-alc-6206-saddle-tan-cover",
            "sort_order": 1,
            "image": "/uploads/alcantara/cases/c-alc-6206-seat-center.jpg",
            "alt_en": "Saddle Tan cover applied to a seat center and armrest program",
            "alt_ja": "シートセンターとアームレストに採用した Saddle Tan カバー",
        },
    ],
}

WIDTHS = {
    "product_types": [26, 18, 22, 20, 36, 28, 64, 38, 44],
    "product_type_specs": [24, 12, 18, 18, 16, 24, 40, 30],
    "product_type_certifications": [24, 12, 44, 30],
    "product_type_maintenance": [24, 12, 22, 18, 54, 36],
    "skus": [30, 16, 24, 20, 22, 18, 12, 40, 40, 56, 34, 38, 30, 56, 34, 40],
    "sku_specs": [30, 12, 24, 18, 42, 28],
    "sku_downloads": [30, 12, 24, 20, 48, 30, 44, 14],
    "sku_case_gallery": [30, 12, 44, 42, 30],
}

NOTES = {
    "product_types": "每行代表一个产品类型，图片字段建议填 /uploads/... 站内路径。",
    "product_type_specs": "定义每个产品类型默认展示的规格字段，站点详情页会按 sort_order 展示。",
    "product_type_certifications": "用于产品类型级认证/卖点展示，可放环保、耐久、法规等短句。",
    "product_type_maintenance": "用于详情页维护说明模块，建议一行一个标题块。",
    "skus": "每行代表一个色号或具体 SKU；image 与 swatch_image 均可使用 /uploads/... 本地路径。",
    "sku_specs": "SKU 级规格覆盖，用于具体色号/品番差异化参数。",
    "sku_downloads": "资料类型仅支持 catalog / technical / care；本地 PDF 可放在 public/catalogs。",
    "sku_case_gallery": "案例图用于详情页底部图库，一行一张，建议放在 /uploads/alcantara/cases/。",
}


def style_sheet(sheet, headers: list[str], rows: list[dict[str, object]]) -> None:
    total_columns = len(headers)
    header_row = 1
    data_start_row = 2

    for column_index, header in enumerate(headers, start=1):
        cell = sheet.cell(row=header_row, column=column_index)
        cell.value = header
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER

    sheet["A1"].comment = Comment(NOTES[sheet.title], "Codex")
    sheet.row_dimensions[1].height = 22

    for row_index, row in enumerate(rows, start=data_start_row):
        for column_index, header in enumerate(headers, start=1):
            cell = sheet.cell(row=row_index, column=column_index)
            cell.value = row.get(header, "")
            cell.font = BODY_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(
                horizontal="left",
                vertical="top",
                wrap_text=len(str(cell.value or "")) > 28,
            )
            if row_index % 2 == 1:
                cell.fill = SUBTLE_FILL

    if rows:
        end_row = data_start_row + len(rows) - 1
        end_column_letter = get_column_letter(total_columns)
        table = Table(displayName=f"{sheet.title.replace('-', '_')}_table", ref=f"A{header_row}:{end_column_letter}{end_row}")
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        sheet.add_table(table)
        sheet.auto_filter.ref = f"A{header_row}:{end_column_letter}{end_row}"

    for idx, width in enumerate(WIDTHS[sheet.title], start=1):
        sheet.column_dimensions[get_column_letter(idx)].width = width

    sheet.freeze_panes = "A2"
    sheet.sheet_view.showGridLines = False


def build_workbook() -> Workbook:
    workbook = Workbook()
    workbook.remove(workbook.active)

    for sheet_name, headers in SHEETS.items():
        sheet = workbook.create_sheet(sheet_name)
        style_sheet(sheet, headers, DATA[sheet_name])

    return workbook


def main() -> None:
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    workbook = build_workbook()
    workbook.save(OUTPUT_PATH)
    print(f"Wrote sample workbook to {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
