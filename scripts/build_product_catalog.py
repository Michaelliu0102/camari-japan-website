#!/usr/bin/env python3

from __future__ import annotations

import argparse
import json
from collections import defaultdict
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook


SHEETS = {
    "product_types": [
        "product_type_slug",
        "material_slug",
        "name_en",
        "name_ja",
        "summary_en",
        "summary_ja",
        "seo_title_en",
        "seo_title_ja",
        "seo_description_en",
        "seo_description_ja",
        "seo_image",
        "product_code",
    ],
    "product_type_specs": [
        "product_type_slug",
        "sort_order",
        "spec_key",
        "label_en",
        "label_ja",
        "aliases",
        "default_value_en",
        "default_value_ja",
    ],
    "product_type_certifications": [
        "product_type_slug",
        "sort_order",
        "text_en",
        "text_ja",
    ],
    "product_type_maintenance": [
        "product_type_slug",
        "sort_order",
        "title_en",
        "title_ja",
        "description_en",
        "description_ja",
    ],
    "skus": [
        "sku_slug",
        "material_slug",
        "product_type_slug",
        "code",
        "color_name_en",
        "color_name_ja",
        "hex",
        "image",
        "swatch_image",
        "summary_en",
        "summary_ja",
        "seo_title_en",
        "seo_title_ja",
        "seo_description_en",
        "seo_description_ja",
        "seo_image",
    ],
    "sku_specs": [
        "sku_slug",
        "sort_order",
        "label_en",
        "label_ja",
        "value_en",
        "value_ja",
    ],
    "product_type_downloads": [
        "product_type_slug",
        "sort_order",
        "title_en",
        "title_ja",
        "description_en",
        "description_ja",
        "href",
        "type",
    ],
    "sku_case_gallery": [
        "sku_slug",
        "sort_order",
        "image",
        "alt_en",
        "alt_ja",
    ],
}


TEMPLATE_ROWS = {
    "product_types": [
        {
            "product_type_slug": "alcantara-panel",
            "material_slug": "alcantara",
            "name_en": "Alcantara Panel",
            "name_ja": "Alcantara パネル",
            "summary_en": "Alcantara panel for automotive door panels, dashboards, and headliners. Italian microfibre with soft-touch finish, UV-stable, and carbon neutral.",
            "summary_ja": "自動車のドアパネル、ダッシュボード、ヘッドライナー向け Alcantara パネル。ソフトタッチ仕上げ、紫外線安定性、カーボンニュートラルなイタリア製マイクロファイバー。",
            "seo_title_en": "Alcantara Panel for Automotive Interiors | CAMARI JAPAN",
            "seo_title_ja": "自動車内装向け Alcantara パネル | CAMARI JAPAN",
            "seo_description_en": "Alcantara panel for automotive door panels, dashboards, and headliners. Italian microfibre with soft-touch finish, UV-stable, and carbon neutral with fire retardant option. View specs and downloads.",
            "seo_description_ja": "自動車のドアパネル、ダッシュボード、ヘッドライナー向け Alcantara パネル。ソフトタッチ仕上げ、紫外線安定性、カーボンニュートラル、難燃オプションあり。仕様と資料をご覧ください。",
            "seo_image": "/uploads/alcantara/panel/alcantara-panel-hero.jpg",
        }
    ],
    "product_type_specs": [
        {
            "product_type_slug": "alcantara-panel",
            "sort_order": 1,
            "spec_key": "thickness",
            "label_en": "THICKNESS",
            "label_ja": "厚み",
            "aliases": "thickness",
            "default_value_en": "ON REQUEST",
            "default_value_ja": "お問い合わせください",
        }
    ],
    "product_type_certifications": [
        {
            "product_type_slug": "alcantara-panel",
            "sort_order": 1,
            "text_en": "Carbon neutral production program",
            "text_ja": "カーボンニュートラル生産プログラム",
        }
    ],
    "product_type_maintenance": [
        {
            "product_type_slug": "alcantara-panel",
            "sort_order": 1,
            "title_en": "Care and Maintenance Guide",
            "title_ja": "ケア・メンテナンスガイド",
            "description_en": "Use and maintenance guidance for installed surfaces.",
            "description_ja": "施工後の使用とメンテナンスのガイド。",
        }
    ],
    "skus": [
        {
            "sku_slug": "c-alc-4991-shadow-black",
            "material_slug": "alcantara",
            "product_type_slug": "alcantara-panel",
            "code": "C-ALC-4991",
            "color_name_en": "Shadow Black",
            "color_name_ja": "シャドウブラック",
            "hex": "#1A1A1A",
            "image": "/uploads/alcantara/panel/c-alc-4991-main.jpg",
            "swatch_image": "/uploads/alcantara/swatches/c-alc-4991-swatch.jpg",
            "summary_en": "A deep charcoal tone for refined interiors.",
            "summary_ja": "上質な空間に向けたディープチャコールトーン。",
            "seo_title_en": "C-ALC-4991 Shadow Black | CAMARI JAPAN",
            "seo_title_ja": "C-ALC-4991 シャドウブラック | CAMARI JAPAN",
            "seo_description_en": "View Shadow Black Alcantara specifications and downloads.",
            "seo_description_ja": "シャドウブラック Alcantara の仕様と資料をご覧ください。",
            "seo_image": "/uploads/alcantara/panel/c-alc-4991-main.jpg",
        }
    ],
    "sku_specs": [
        {
            "sku_slug": "c-alc-4991-shadow-black",
            "sort_order": 1,
            "label_en": "Thickness",
            "label_ja": "厚み",
            "value_en": "0.95 mm",
            "value_ja": "0.95 mm",
        }
    ],
    "product_type_downloads": [
        {
            "product_type_slug": "alcantara-panel",
            "sort_order": 1,
            "title_en": "Alcantara Technical Sheet",
            "title_ja": "Alcantara 技術資料",
            "description_en": "Composition, width, care, and performance notes.",
            "description_ja": "組成、幅、ケア、性能情報。",
            "href": "https://example.com/alcantara-technical-sheet.pdf",
            "type": "technical",
        }
    ],
    "sku_case_gallery": [
        {
            "sku_slug": "c-alc-4991-shadow-black",
            "sort_order": 1,
            "image": "/uploads/alcantara/cases/c-alc-4991-installed.jpg",
            "alt_en": "Installed case image",
            "alt_ja": "施工事例画像",
        }
    ],
}


def localized(en: str, ja: str) -> dict[str, str]:
    return {"en": en or "", "ja": ja or ""}


def normalize_image_path(value: str) -> str:
    """Strip absolute local path prefix to make it relative to public/."""
    if not value:
        return value
    # Strip /Users/.../public or /Users/.../Website/public prefix
    import re
    value = re.sub(r'^.*?/public(?=/|$)', '', value)
    return value or ""


def read_rows(workbook_path: Path) -> dict[str, list[dict[str, Any]]]:
    workbook = load_workbook(workbook_path)
    rows_by_sheet: dict[str, list[dict[str, Any]]] = {}

    for sheet_name, headers in SHEETS.items():
      if sheet_name not in workbook.sheetnames:
          raise ValueError(f"Missing worksheet: {sheet_name}")

      sheet = workbook[sheet_name]
      header_row = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
      if header_row != headers:
          raise ValueError(f"Worksheet {sheet_name} headers do not match template")

      rows: list[dict[str, Any]] = []
      for values in sheet.iter_rows(min_row=2, values_only=True):
          row = {header: ("" if value is None else str(value).strip()) for header, value in zip(headers, values)}
          if any(row.values()):
              rows.append(row)
      rows_by_sheet[sheet_name] = rows

    return rows_by_sheet


def sort_key(row: dict[str, Any]) -> tuple[int, str]:
    raw = row.get("sort_order", "")
    try:
        order = int(raw)
    except ValueError:
        order = 9999
    return (order, json.dumps(row, sort_keys=True))


def split_aliases(value: str) -> list[str]:
    return [alias.strip() for alias in value.split("|") if alias.strip()]


def build_catalog(rows_by_sheet: dict[str, list[dict[str, Any]]]) -> dict[str, Any]:
    specs_by_type: dict[str, list[dict[str, Any]]] = defaultdict(list)
    certs_by_type: dict[str, list[dict[str, Any]]] = defaultdict(list)
    maintenance_by_type: dict[str, list[dict[str, Any]]] = defaultdict(list)
    specs_by_sku: dict[str, list[dict[str, Any]]] = defaultdict(list)
    downloads_by_type: dict[str, list[dict[str, Any]]] = defaultdict(list)
    case_gallery_by_sku: dict[str, list[dict[str, Any]]] = defaultdict(list)

    for row in sorted(rows_by_sheet["product_type_specs"], key=sort_key):
        specs_by_type[row["product_type_slug"]].append(
            {
                "key": row["spec_key"],
                "label": localized(row["label_en"], row["label_ja"]),
                "aliases": split_aliases(row["aliases"]),
                "defaultValue": localized(row["default_value_en"], row["default_value_ja"]),
            }
        )

    for row in sorted(rows_by_sheet["product_type_certifications"], key=sort_key):
        certs_by_type[row["product_type_slug"]].append(localized(row["text_en"], row["text_ja"]))

    for row in sorted(rows_by_sheet["product_type_maintenance"], key=sort_key):
        maintenance_by_type[row["product_type_slug"]].append(
            {
                "title": localized(row["title_en"], row["title_ja"]),
                "description": localized(row["description_en"], row["description_ja"]),
            }
        )

    for row in sorted(rows_by_sheet["sku_specs"], key=sort_key):
        specs_by_sku[row["sku_slug"]].append(
            {
                "label": localized(row["label_en"], row["label_ja"]),
                "value": localized(row["value_en"], row["value_ja"]),
            }
        )

    for row in sorted(rows_by_sheet["product_type_downloads"], key=sort_key):
        download_type = row["type"].lower()
        if download_type not in {"catalog", "technical", "care"}:
            raise ValueError(f"Unsupported download type for {row['product_type_slug']}: {row['type']}")
        downloads_by_type[row["product_type_slug"]].append(
            {
                "title": localized(row["title_en"], row["title_ja"]),
                "description": localized(row["description_en"], row["description_ja"]),
                "href": row["href"],
                "type": download_type,
            }
        )

    for row in sorted(rows_by_sheet["sku_case_gallery"], key=sort_key):
        case_gallery_by_sku[row["sku_slug"]].append(
            {
                "image": normalize_image_path(row["image"]),
                "alt": localized(row["alt_en"], row["alt_ja"]),
            }
        )

    product_types: list[dict[str, Any]] = []
    known_product_types: set[str] = set()
    for row in rows_by_sheet["product_types"]:
        slug = row["product_type_slug"]
        if not slug or not row.get("name_en"):
            continue  # skip template rows
        known_product_types.add(slug)
        product_types.append(
            {
                "slug": slug,
                "materialSlug": row["material_slug"],
                "name": localized(row["name_en"], row["name_ja"]),
                "summary": localized(row["summary_en"], row["summary_ja"]),
                "productCode": row.get("product_code", ""),
                "downloads": downloads_by_type.get(slug, []),
                "specTemplate": specs_by_type.get(slug, []),
                "certifications": certs_by_type.get(slug, []),
                "maintenance": maintenance_by_type.get(slug, []),
                "seo": {
                    "title": localized(row["seo_title_en"], row["seo_title_ja"]),
                    "description": localized(row["seo_description_en"], row["seo_description_ja"]),
                    "image": normalize_image_path(row["seo_image"]),
                },
            }
        )

    skus: list[dict[str, Any]] = []
    for row in rows_by_sheet["skus"]:
        if not row.get("sku_slug"):
            continue  # skip empty rows
        product_type_slug = row["product_type_slug"]
        if product_type_slug not in known_product_types:
            raise ValueError(f"SKU {row['sku_slug']} references missing product type: {product_type_slug}")

        sku = {
            "slug": row["sku_slug"],
            "materialSlug": row["material_slug"],
            "productTypeSlug": product_type_slug,
            "code": row["code"],
            "colorName": localized(row["color_name_en"], row["color_name_ja"]),
            "hex": row["hex"],
            "image": normalize_image_path(row["image"]),
            "swatchImage": normalize_image_path(row["swatch_image"]) or None,
            "caseGallery": case_gallery_by_sku.get(row["sku_slug"], []),
            "summary": localized(row["summary_en"], row["summary_ja"]),
            "specs": specs_by_sku.get(row["sku_slug"], []),
            "certifications": [],
            "seo": {
                "title": localized(row["seo_title_en"], row["seo_title_ja"]),
                "description": localized(row["seo_description_en"], row["seo_description_ja"]),
                "image": normalize_image_path(row["seo_image"] or row["image"]),
            },
        }
        if not sku["swatchImage"]:
            del sku["swatchImage"]
        skus.append(sku)

    return {"productTypes": product_types, "skus": skus}


def write_catalog_json(catalog: dict[str, Any], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(catalog, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def write_template(output_path: Path) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)

    for sheet_name, headers in SHEETS.items():
        sheet = workbook.create_sheet(sheet_name)
        sheet.append(headers)
        for row in TEMPLATE_ROWS.get(sheet_name, []):
            sheet.append([row.get(header, "") for header in headers])

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Build product type and SKU fixture data from an Excel workbook."
    )
    parser.add_argument(
        "workbooks",
        nargs="*",
        type=Path,
        help="One or more .xlsx workbooks, or a directory containing .xlsx files."
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=Path("src/data/product-catalog.generated.json"),
        help="Where to write the compiled JSON catalog."
    )
    parser.add_argument(
        "--write-template",
        type=Path,
        help="Write a blank template workbook to this path and exit."
    )
    return parser.parse_args()


def resolve_workbooks(paths: list[Path]) -> list[Path]:
    """Resolve a list of paths to .xlsx files. Directories are scanned recursively. Skips temp/lock files."""
    result: list[Path] = []
    for path in paths:
        if path.is_dir():
            for f in sorted(path.rglob("*.xlsx")):
                if not f.name.startswith("~$"):
                    result.append(f)
        elif path.suffix.lower() == ".xlsx" and not path.name.startswith("~$"):
            result.append(path)
    return result


def merge_catalogs(catalogs: list[dict[str, Any]]) -> dict[str, Any]:
    """Merge multiple catalog dicts into one, deduplicating by slug."""
    all_product_types: dict[str, dict[str, Any]] = {}
    all_skus: dict[str, dict[str, Any]] = {}

    for catalog in catalogs:
        for pt in catalog.get("productTypes", []):
            all_product_types[pt["slug"]] = pt
        for sku in catalog.get("skus", []):
            all_skus[sku["slug"]] = sku

    return {
        "productTypes": list(all_product_types.values()),
        "skus": list(all_skus.values()),
    }


def main() -> int:
    args = parse_args()

    if args.write_template:
        write_template(args.write_template)
        print(f"Template workbook written to {args.write_template}")
        return 0

    workbooks = resolve_workbooks(args.workbooks)
    if not workbooks:
        raise SystemExit("Provide at least one workbook path, or use --write-template.")

    catalogs = []
    for wb_path in workbooks:
        print(f"Reading {wb_path}...")
        catalogs.append(build_catalog(read_rows(wb_path)))

    merged = merge_catalogs(catalogs) if len(catalogs) > 1 else catalogs[0]
    write_catalog_json(merged, args.output)
    print(f"\nGenerated catalog JSON at {args.output}")
    print(f"Product types: {len(merged['productTypes'])}")
    print(f"SKUs: {len(merged['skus'])}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
